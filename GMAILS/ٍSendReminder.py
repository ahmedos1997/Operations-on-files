# الرجاء مراجعة هذا الدرس ضروري
#   نريد ارسال رسائل بريد للتذكير بمستحقات الاعضاء الذين لم يدفعو وبياناتهم مسجلة في ملف اكسل
import openpyxl,smtplib
from pathlib import Path

#connect to excel file

file = openpyxl.load_workbook(Path.home() / Path('G:\python cource\project\GMAILS\members.xlsx'))
sheets = file.sheetnames
sheet = file['Sheet1']
lastcol = sheet.max_column
lastmonth = sheet.cell(row=1, column=lastcol).value
print(lastmonth)

# find unpaid members

unpaid = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=sheet.max_column)
    if payment != 'paid':
        name = sheet.cell(row=r , column=1)
        email = sheet.cell(row=r , column=2)
        unpaid[name] = email

# send email for unpaid member

send = smtplib.SMTP('smtp.gmail.com', 587)
send.starttls()
sender_email = 'ahmedelmadani1997@gmail.com'
password = input(str('Enter your password'))
send.login(sender_email,password)

for name, email in unpaid.items():
    body = """ subject : %s dues unpaid.\ndear %s,\n records show that you are not paid last month for %s
    please make this payment as soon .thank you 
    """ % (lastmonth, name , lastmonth)
    print('sending email for %s ' %email)
    sendstatus = send.sendmail(sender_email, email, body)
    if sendstatus != {}:
        print('there are some problem in send email to %s %s' % (email, sendstatus))
send.quit()











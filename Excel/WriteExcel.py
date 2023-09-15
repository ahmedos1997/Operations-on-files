#نستخدم  ()Workbook لإنشاء  ملف إكسل.

import openpyxl
from pathlib import Path

# create excel file

exfile = openpyxl.Workbook()
print(exfile.sheetnames) # للتاكد من انه جهز ملف اكسل
# يجب حفظ الملف ليظهر في الجهاز وذلك كالاتي

# لتغير الاسم الافتراضي لاول ملف داخل الاكسل
exsheet = exfile.active               # كما نذكر ان active هو اول ملف في الاكسل
exsheet.title = 'sheetone'            # title هو العنوان

# انشاء ورقة جديدة داخل الملف
# يمكنك إنشاء ورقة جديدة من خلال الدالة ()create_chartsheet.

exfile.create_sheet()
exfile.create_sheet()
exfile.create_sheet(index=1 , title='sheettow')                # هنا حددنا مكان الملف عن طريق index
exfile.create_sheet(index=2 , title='sheetthird')
# طريقة حذف ملف
del exfile['sheetthird']

# طريقة الكتابة على ملف داخل الاكسل
sheet = exfile['sheetone']                # حددنا اي ملف نريد الكتباة عليه
sheet['A1'] = 'hello, world'               # حددنا في اي عمود سوف نكتب و حددنا ما نريد كتابته
print(sheet['A1'].value)                       # عرضنا ما كتبناه


exfile.save(filename= Path.home() / Path('D:/python cource','project','Excel','example2.xlsx')) # وهكذا تم حفظ الملف

print('-' * 200)






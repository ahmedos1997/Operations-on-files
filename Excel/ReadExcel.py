# عمليات على ملفات الاكسل
#تمكّنك الدالة ()load_workbook من فتح ملف إكسل.
import openpyxl
from pathlib import Path
exelfile = openpyxl.load_workbook(Path.home() / Path('D:/python cource','project','Excel','example.xlsx'))
print(exelfile.sheetnames) # طباعة اسماء الملفات داخل ملف الاكسل

sheet1 = exelfile['Sheet1'] # يجب مراعات الحروف الكبيرة و الصغير في اسم الملف
print(sheet1.title)  # الدالة title  لاستخراج الاسم كعنوان


activesheet = exelfile.active   # و الدالة active تستخدم لمعرفة الملف الاول في الاكسل (الملف النشط)
print(activesheet.title)
print('-'*100)

print(sheet1['A1'].value) #يمكن قراءة الخلية من خلال الخاصية value
print(sheet1['B1'].value)
print(sheet1['C1'].value)

print(sheet1['C1'].row) # لمعرفة السطر
print(sheet1['C1'].column) # لمعرفة العامود

print(sheet1['C1'].coordinate)

print('-'*100)


print(sheet1.cell(row=1, column=2).value) # هنا قلنا للبرنامج اجلب لنا القيمة الموجودة في الصف الاول العاكود الثاني

# مثال لعرض جميع الاسماء في الملف شييت 1

print('-'*100)
for i in range(1, 7) :
    print(i, sheet1.cell(row=i, column=1).value) # يمكن قراءة الخلية أيضًا من خلال الدالة ()cell.

    print('-' * 100)

for i in range (1, 7) :
    print(i, sheet1.cell(row=i , column=2).value)

print('-' * 100)
# تمرين لكشف اسماء الموظفين و مرتباتهم ومجموع مرتباتهم

total = 0
for i in range(1, 7) :
    print(sheet1.cell(row=i, column=1).value , sheet1.cell(row=i, column=2).value)

    total += sheet1.cell(row=i, column=2).value

print(f'the total of your salary is {total}')

#تستخدم الخاصية max_row لمعرفة عدد أسطر الورقة والخاصية max_column لمعرفة عدد أعمدة الورقة.
print(sheet1.max_row) # طباعة كل الاسطر
print(sheet1.max_column) # طباعة كل الاعمدة

print('-' * 100)

# مثال في حالة لم نكن نعلم عدد الصفوف او الاعمدة
# سنعيد مثال عرض اسماء الموظفين

for i in range(1,sheet1.max_row+1) : # زدنا واحد لان البرنامج ياخذ الرقم القبل الاخير
    print(i, sheet1.cell(row=i, column=1).value) # اسماء جميع الموظفين
    print('-' * 100)


